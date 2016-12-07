import argparse

from lxml import html, etree

import csv
import xlrd
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import jinja2
import os
import json
import uuid

CWD = os.path.dirname(os.path.abspath(__file__))
LOG_FILE = None

templateLoader = jinja2.FileSystemLoader( searchpath=CWD+'/templates/' )

templateEnv = jinja2.Environment( loader=templateLoader )


def logging_info(msg, level='INFO'):
    """ We're using JSON which is easier to parse """
    global LOG_FILE
    with open(LOG_FILE, 'a') as fh:
        fh.write(json.dumps({'msg': msg, 'level': level})+'\n')

    print('CONVERTER LOGGING : [%s] %s' % (level, msg))


def get_result_file_name():
    if not os.path.exists(os.path.join(CWD, '_results')):
        os.makedirs(os.path.join(CWD, '_results'))
    filename = os.path.join(CWD, '_results', str(uuid.uuid4()))
    return filename


def write_to_file(content):
    filename = get_result_file_name()
    with open(filename, 'w') as result:
        result.write(content)

    return filename


def check_extension(filename, extensions):
    name, file_extension = os.path.splitext(filename)
    print name
    print file_extension
    return file_extension in extensions


def generate_google_manufacturer_xml(input_file):
    available_extensions = ['.csv']
    items = []
    context = {}

    if not check_extension(input_file, available_extensions):
        logging_info('The file extension should be %s.'
                     % (','.join(available_extensions)), 'ERROR')
        return

    try:
        with open(input_file, 'rU') as csvfile:
            csvfile.readline()
            reader = csv.reader(csvfile)
            for item in reader:
                tree_description = html.fromstring(item[10])
                tree_bullets = \
                    tree_description.xpath("//*[contains(@id,'feature-bullets')]//ul/"
                                           "li[not(contains(@class,'hidden'))]")
                bullet_points = []
                for bullet in tree_bullets:
                    bullet_points.append(bullet.text_content())

                items.append({
                    'id': item[7],
                    'brand': item[17],
                    'title': item[9],
                    'gtin': item[1],
                    'mpn': item[7],                                 # This field should be redefined
                    'description': item[12],
                    'bullet_points': bullet_points,
                })
    except Exception as e:
        logging_info(str(e), 'ERROR')
        return

    context['items'] = items

    template = templateEnv.get_template('GoogleManufacturer.html')
    output_content = template.render(context)

    filename = write_to_file(output_content)

    logging_info(filename, 'RESULT_FILE')
    logging_info('google-manufacturer.xml', 'FILE_NAME')


def generate_amazon_to_walmart(input_file):
    available_extensions = ['.xls']

    print input_file

    if not check_extension(input_file, available_extensions):
        logging_info('The file extension should be %s.'
                     % (','.join(available_extensions)), 'ERROR')
        return

    template = templateEnv.get_template('walmart_long_description.html')

    def create_walmart_description(bullet1, bullet2, bullet3, bullet4, bullet5):
        context = {
            "bullet1": bullet1 if bullet1 != "" else None,
            "bullet2": bullet2 if bullet2 != "" else None,
            "bullet3": bullet3 if bullet3 != "" else None,
            "bullet4": bullet4 if bullet4 != "" else None,
            "bullet5": bullet5 if bullet5 != "" else None,
        }
        return template.render(context)

    try:
        wb = xlrd.open_workbook(filename=input_file)
    except Exception as e:
        logging_info(str(e), 'ERROR')
        return

    item_sheet = wb.sheet_by_name('Item_Sheet')

    items = []
    # Read data from amazone form
    for idx, row in enumerate(item_sheet.get_rows()):
        if idx > 3:
            data = {
                "tool_id": row[4].value,
                "title": row[6].value,
                "short_desc": row[30].value,
                "long_desc": create_walmart_description(
                    row[31].value, row[32].value, row[33].value, row[34].value, row[35].value),
                "ingredients": row[37].value,
                "directions": row[38].value,
                "warnings": row[38].value,
            }
            if row[16].value == "UPC":
                data["upc"] = row[17].value
                data["gstn"] = ""
            elif row[16].value == "GSTN":
                data["upc"] = ""
                data["gstn"] = row[17].value
            else:
                data["upc"] = ""
                data["gstn"] = ""

            items.append(data)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"

    # Create header
    ws1.append(["GTIN-14",
                "UPC",
                "Tool ID",
                "Product Name",
                "Long Description",
                "Shelf Description (Optional)",
                "Short Description",
                "Usage Directions (optional)",
                "Ingredients (optional)",
                "Caution Warnings Allergens (optional)"])
    # write product info
    for item in items:
        ws1.append([item["gstn"],
                    item["upc"],
                    item["tool_id"],
                    item["title"],
                    item["long_desc"],
                    "",
                    item["short_desc"],
                    item["ingredients"],
                    item["directions"],
                    item["warnings"],
                    ])

    filename = get_result_file_name()
    wb.save(filename=filename)

    logging_info(filename, 'RESULT_FILE')
    logging_info('WalmartBulkContentUpload.xlsx', 'FILE_NAME')


def generate_error():
    logging_info('This type of conversion does not exist.', 'ERROR')


def main():
    global LOG_FILE, OUTPUT_DIR, ID

    parser = argparse.ArgumentParser(description='test',
                                     formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument('--input_type', type=str, required=True,
                        help="Enter input type\n"
                             "Defined types\n"
                             " - amazon\n"
                             " - template\n")
    parser.add_argument('--output_type', type=str, required=True,
                        help="Enter output type\n"
                             "Defined types\n"
                             " - walmart\n"
                             " - googlemanufacturer\n")
    parser.add_argument('--input_file', type=str, required=True,
                        help="File to upload")
    parser.add_argument('--mapping_file', type=str, required=True,
                        help="File to map")
    parser.add_argument('--log_file', type=str, required=True,
                        help="filename for output logging")

    namespace = parser.parse_args()


    LOG_FILE = namespace.log_file
    input_file = namespace.input_file
    input_type = namespace.input_type
    output_type = namespace.output_type

    if input_type == 'template' and output_type == 'googlemanufacturer':
        generate_google_manufacturer_xml(input_file)
    elif input_type == 'amazon' and output_type == 'walmart':
        generate_amazon_to_walmart(input_file)
    else:
        generate_error()


# class AmazonToWalmartView(View):
#     def get(self, request, *args, **kwargs):
#         template = loader.get_template('walmart_long_description.html')
#         def create_walmart_description(bullet1, bullet2, bullet3, bullet4, bullet5):
#             context = {
#                 "bullet1": bullet1 if bullet1 != "" else None,
#                 "bullet2": bullet2 if bullet2 != "" else None,
#                 "bullet3": bullet3 if bullet3 != "" else None,
#                 "bullet4": bullet4 if bullet4 != "" else None,
#                 "bullet5": bullet5 if bullet5 != "" else None,
#             }
#             return template.render(context, request)
#
#         filename = settings.BASE_DIR + \
#                    settings.MEDIA_URL + \
#                    'amazon_to_walmart/Amazon-HPC-Template-Export-2016-10-25(22-26-50).xls'
#
#         wb = xlrd.open_workbook(filename = filename)
#         item_sheet = wb.sheet_by_name('Item_Sheet')
#
#         items = []
#         # Read data from amazone form
#         for idx, row in enumerate(item_sheet.get_rows()):
#             if idx > 3:
#                 data = {
#                     "tool_id": row[4].value,
#                     "title": row[6].value,
#                     "short_desc": row[30].value,
#                     "long_desc": create_walmart_description(
#                         row[31].value, row[32].value, row[33].value, row[34].value, row[35].value),
#                     "ingredients": row[37].value,
#                     "directions": row[38].value,
#                     "warnings": row[38].value,
#                 }
#                 if row[16].value == "UPC":
#                     data["upc"] = row[17].value
#                     data["gstn"] = ""
#                 elif row[16].value == "GSTN":
#                     data["upc"] = ""
#                     data["gstn"] = row[17].value
#                 else:
#                     data["upc"] = ""
#                     data["gstn"] = ""
#
#                 items.append(data)
#
#         wb = Workbook()
#         dest_filename = 'empty_book.xlsx'
#         ws1 = wb.active
#         ws1.title = "Sheet1"
#
#         # Create header
#         ws1.append(["GTIN-14",
#                     "UPC",
#                     "Tool ID",
#                     "Product Name",
#                     "Long Description",
#                     "Shelf Description (Optional)",
#                     "Short Description",
#                     "Usage Directions (optional)",
#                     "Ingredients (optional)",
#                     "Caution Warnings Allergens (optional)"])
#         # write product info
#         for item in items:
#             ws1.append([item["gstn"],
#                         item["upc"],
#                         item["tool_id"],
#                         item["title"],
#                         item["long_desc"],
#                         "",
#                         item["short_desc"],
#                         item["ingredients"],
#                         item["directions"],
#                         item["warnings"],])
#
#         temp_filename = filename = settings.BASE_DIR + \
#                    settings.MEDIA_URL + \
#                    "tmp/" + str(uuid.uuid4()) + ".xlsx"
#         wb.save(filename = temp_filename)
#
#         fp = open(temp_filename, 'rb')
#         response = HttpResponse(fp.read())
#         fp.close()
#         response['Content-Type'] = 'application/vnd.ms-excel'
#         # response['Content-Type'] = 'application/force-download'
#         response['Content-Disposition'] = 'attachment; filename=WalmartBulkContentUpload.xlsx'
#         return response


if __name__ == '__main__':
    main()
